#!/usr/bin/env python3
"""
Excel (XLS/XLSX) to Markdown Converter for Common Data Set
Extracts university data from Excel files and converts to formatted markdown.
"""

import os
import sys
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import xlrd
import re


class ExcelToMarkdownConverter:
    """Convert Excel files to Markdown format."""

    def __init__(self, output_dir: str = "documents"):
        """
        Initialize the converter.

        Args:
            output_dir: Directory to save markdown files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def process_xls(self, file_path: str) -> str:
        """
        Process an .xls file (Excel 97-2003).

        Args:
            file_path: Path to the .xls file

        Returns:
            Markdown content
        """
        print(f"Processing XLS file: {file_path}")
        workbook = xlrd.open_workbook(file_path)
        markdown_sections = []

        # Get university name from filename
        uni_name = Path(file_path).stem
        markdown_sections.append(f"# {uni_name} Common Data Set 2024-2025\n")

        # Process each sheet
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            print(f"  Processing sheet: {sheet.name} ({sheet.nrows} rows, {sheet.ncols} cols)")

            section_md = self._process_xls_sheet(sheet)
            if section_md:
                markdown_sections.append(section_md)

        return "\n".join(markdown_sections)

    def process_xlsx(self, file_path: str) -> str:
        """
        Process an .xlsx file (Excel 2007+).

        Args:
            file_path: Path to the .xlsx file

        Returns:
            Markdown content
        """
        print(f"Processing XLSX file: {file_path}")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        markdown_sections = []

        # Get university name from filename
        uni_name = Path(file_path).stem
        markdown_sections.append(f"# {uni_name} Common Data Set 2024-2025\n")

        # Process each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"  Processing sheet: {sheet_name} ({sheet.max_row} rows, {sheet.max_column} cols)")

            section_md = self._process_xlsx_sheet(sheet)
            if section_md:
                markdown_sections.append(section_md)

        return "\n".join(markdown_sections)

    def _process_xls_sheet(self, sheet: xlrd.sheet.Sheet) -> str:
        """Process a single XLS sheet."""
        lines = []
        section_name = sheet.name

        # Add section header
        if section_name:
            lines.append(f"\n## {section_name}\n")

        # Extract all data into a grid
        data_grid = []
        for row_idx in range(sheet.nrows):
            row_data = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                value = self._format_cell_value_xls(cell)
                row_data.append(value)
            data_grid.append(row_data)

        # Process grid into markdown
        markdown = self._grid_to_markdown(data_grid)
        lines.append(markdown)

        return "\n".join(lines)

    def _process_xlsx_sheet(self, sheet: Worksheet) -> str:
        """Process a single XLSX sheet."""
        lines = []
        section_name = sheet.title

        # Add section header
        if section_name:
            lines.append(f"\n## {section_name}\n")

        # Extract all data into a grid
        data_grid = []
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                value = self._format_cell_value_xlsx(cell)
                row_data.append(value)
            data_grid.append(row_data)

        # Process grid into markdown
        markdown = self._grid_to_markdown(data_grid)
        lines.append(markdown)

        return "\n".join(lines)

    def _format_cell_value_xls(self, cell: xlrd.sheet.Cell) -> str:
        """Format XLS cell value as string."""
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return ""
        elif cell.ctype == xlrd.XL_CELL_TEXT:
            return str(cell.value).strip()
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            # Format numbers without unnecessary decimals
            if cell.value == int(cell.value):
                return str(int(cell.value))
            return str(cell.value)
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "✓" if cell.value else ""
        else:
            return str(cell.value).strip()

    def _format_cell_value_xlsx(self, cell: Cell) -> str:
        """Format XLSX cell value as string."""
        if cell.value is None:
            return ""
        elif isinstance(cell.value, bool):
            return "✓" if cell.value else ""
        elif isinstance(cell.value, (int, float)):
            # Format numbers without unnecessary decimals
            if isinstance(cell.value, float) and cell.value == int(cell.value):
                return str(int(cell.value))
            return str(cell.value)
        else:
            return str(cell.value).strip()

    def _grid_to_markdown(self, grid: List[List[str]]) -> str:
        """
        Convert a data grid to markdown format.
        Intelligently detects tables, lists, and key-value pairs.
        """
        if not grid or all(not any(row) for row in grid):
            return ""

        lines = []
        i = 0

        while i < len(grid):
            row = grid[i]

            # Skip empty rows
            if not any(row):
                lines.append("")
                i += 1
                continue

            # Check if this looks like a section header (single cell with text)
            if self._is_section_header(row):
                header_text = next(cell for cell in row if cell)
                # Determine header level
                if self._is_main_section(header_text):
                    lines.append(f"### {header_text}\n")
                else:
                    lines.append(f"#### {header_text}\n")
                i += 1
                continue

            # Check if this is the start of a table
            table_rows = self._extract_table(grid, i)
            if table_rows and len(table_rows) > 1:
                # Format as markdown table
                table_md = self._format_markdown_table(table_rows)
                lines.append(table_md)
                lines.append("")
                i += len(table_rows)
                continue

            # Check if this is a key-value pair
            if self._is_key_value_pair(row):
                kv_md = self._format_key_value(row)
                lines.append(kv_md)
                i += 1
                continue

            # Check if this is a list item
            if self._is_list_item(row):
                list_md = self._format_list_item(row)
                lines.append(list_md)
                i += 1
                continue

            # Default: just join non-empty cells
            non_empty = [cell for cell in row if cell]
            if non_empty:
                lines.append(" | ".join(non_empty))

            i += 1

        return "\n".join(lines)

    def _is_section_header(self, row: List[str]) -> bool:
        """Check if row is a section header (single merged cell or mostly empty)."""
        non_empty = [cell for cell in row if cell]
        return len(non_empty) == 1

    def _is_main_section(self, text: str) -> bool:
        """Check if text is a main section (like 'A. General Information')."""
        # Main sections typically start with a letter and period
        return bool(re.match(r'^[A-Z]\d*\.?\s+', text))

    def _is_key_value_pair(self, row: List[str]) -> bool:
        """Check if row is a key-value pair (2 non-empty cells)."""
        non_empty = [cell for cell in row if cell]
        return len(non_empty) == 2

    def _is_list_item(self, row: List[str]) -> bool:
        """Check if row is a list item (single cell, not a header)."""
        non_empty = [cell for cell in row if cell]
        if len(non_empty) != 1:
            return False
        text = non_empty[0]
        # List items are usually short and don't look like section headers
        return len(text) < 100 and not self._is_main_section(text)

    def _extract_table(self, grid: List[List[str]], start_idx: int) -> Optional[List[List[str]]]:
        """
        Extract a table starting from start_idx.
        Returns None if not a table.
        """
        if start_idx >= len(grid):
            return None

        # Check if first row could be a header (multiple non-empty cells)
        first_row = grid[start_idx]
        non_empty_count = sum(1 for cell in first_row if cell)

        if non_empty_count < 2:
            return None

        # Determine number of columns
        num_cols = len(first_row)

        # Collect rows that have the same structure
        table_rows = [first_row]
        for i in range(start_idx + 1, len(grid)):
            row = grid[i]

            # Stop if we hit an empty row or row with different structure
            if not any(row):
                break

            # Check if row has cells in similar positions
            if len(row) == num_cols:
                table_rows.append(row)
            else:
                break

            # Stop after a reasonable number of rows
            if len(table_rows) >= 50:
                break

        # Need at least 2 rows for a valid table
        if len(table_rows) < 2:
            return None

        return table_rows

    def _format_markdown_table(self, rows: List[List[str]]) -> str:
        """Format rows as a markdown table."""
        if not rows:
            return ""

        # Calculate max columns
        max_cols = max(len(row) for row in rows)

        # Normalize all rows to have the same number of columns
        normalized_rows = []
        for row in rows:
            normalized = row + [""] * (max_cols - len(row))
            normalized_rows.append(normalized)

        # Build table
        lines = []

        # Header row
        header = "| " + " | ".join(normalized_rows[0]) + " |"
        lines.append(header)

        # Separator
        separator = "|" + "|".join(["-------"] * max_cols) + "|"
        lines.append(separator)

        # Data rows
        for row in normalized_rows[1:]:
            data_row = "| " + " | ".join(row) + " |"
            lines.append(data_row)

        return "\n".join(lines)

    def _format_key_value(self, row: List[str]) -> str:
        """Format a key-value pair."""
        non_empty = [cell for cell in row if cell]
        if len(non_empty) == 2:
            return f"| {non_empty[0]} | {non_empty[1]} |"
        return " | ".join(non_empty)

    def _format_list_item(self, row: List[str]) -> str:
        """Format a list item."""
        non_empty = [cell for cell in row if cell]
        if non_empty:
            return f"- {non_empty[0]}"
        return ""

    def process_file(self, file_path: str, skip_if_processed: bool = False) -> bool:
        """
        Process an Excel file and save as markdown.

        Args:
            file_path: Path to the Excel file
            skip_if_processed: If True, skip files that have already been processed

        Returns:
            True if processed successfully, False otherwise
        """
        file_path_obj = Path(file_path)

        if not file_path_obj.exists():
            print(f"Error: File not found: {file_path}")
            return False

        # Create tracking file paths
        processed_marker = self.output_dir / f".{file_path_obj.stem}.processed"
        failed_marker = self.output_dir / f".{file_path_obj.stem}.failed"
        output_file = self.output_dir / f"{file_path_obj.stem}.md"

        # Skip if already processed successfully
        if skip_if_processed and processed_marker.exists() and output_file.exists():
            return None  # Indicates skipped

        # Remove old failed marker if retrying
        if failed_marker.exists():
            failed_marker.unlink()

        # Determine file type and process
        try:
            if file_path_obj.suffix.lower() == '.xls':
                markdown_content = self.process_xls(file_path)
            elif file_path_obj.suffix.lower() in ['.xlsx', '.xlsm']:
                markdown_content = self.process_xlsx(file_path)
            else:
                print(f"Error: Unsupported file format: {file_path_obj.suffix}")
                return False

            # Save markdown file
            with open(output_file, 'w', encoding='utf-8') as f:
                f.write(markdown_content)

            # Create success marker
            with open(processed_marker, 'w') as f:
                f.write(f"Processed on: {file_path_obj.stat().st_mtime}\n")
                f.write(f"Output: {output_file}\n")

            print(f"✓ Saved to: {output_file}\n")
            return True

        except Exception as e:
            print(f"✗ Error processing {file_path}: {e}\n")

            # Create failure marker
            with open(failed_marker, 'w') as f:
                f.write(f"Failed with error: {str(e)}\n")

            import traceback
            traceback.print_exc()
            return False

    def process_directory(self, directory: str = ".") -> None:
        """
        Process all Excel files in a directory.

        Args:
            directory: Directory containing Excel files
        """
        dir_path = Path(directory)
        excel_files = []

        # Find all Excel files
        for pattern in ['*.xls', '*.xlsx', '*.xlsm']:
            excel_files.extend(sorted(dir_path.glob(pattern)))

        if not excel_files:
            print(f"No Excel files found in {directory}")
            return

        print(f"Found {len(excel_files)} Excel file(s)\n")

        # Track processing statistics
        processed_count = 0
        skipped_count = 0
        failed_count = 0

        for file_path in excel_files:
            result = self.process_file(str(file_path), skip_if_processed=True)

            if result is True:
                processed_count += 1
            elif result is False:
                failed_count += 1
            elif result is None:
                print(f"⊙ Skipping {file_path.name} (already processed)")
                skipped_count += 1

        # Print summary
        print("\n" + "="*60)
        print(f"Processing Summary:")
        print(f"  ✓ Processed: {processed_count}")
        print(f"  ⊙ Skipped:   {skipped_count}")
        print(f"  ✗ Failed:    {failed_count}")
        print("="*60)


def main():
    """Main entry point for the script."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Convert Excel (XLS/XLSX) files to Markdown for Common Data Set"
    )
    parser.add_argument(
        'input',
        nargs='?',
        help='Excel file or directory to process (default: current directory)'
    )
    parser.add_argument(
        '-o', '--output-dir',
        default='documents',
        help='Output directory for markdown files (default: documents)'
    )

    args = parser.parse_args()

    converter = ExcelToMarkdownConverter(output_dir=args.output_dir)

    # Process input
    if args.input is None:
        # Process all Excel files in current directory
        converter.process_directory(".")
    elif os.path.isfile(args.input):
        # Process single file
        converter.process_file(args.input)
    elif os.path.isdir(args.input):
        # Process directory
        converter.process_directory(args.input)
    else:
        print(f"Error: {args.input} is not a valid file or directory")
        sys.exit(1)


if __name__ == "__main__":
    main()
